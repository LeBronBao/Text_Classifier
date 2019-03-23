# -*- encoding: utf-8 -*-

from keras import backend as K
from keras.models import Sequential, Model, load_model
from keras.layers.core import Dense, Dropout
from keras.layers import LSTM, Bidirectional, concatenate, Input, Embedding, Add, GRU, BatchNormalization, Flatten
from keras.layers import Conv1D, GlobalMaxPool1D, GlobalAveragePooling1D, Multiply, Dot, Average, Lambda
from keras.activations import softmax
from attention import Attention
from data_preprocess import preprocess_amazon_mobile_data
from result_statistic import read_test_data
import openpyxl as xl
import matplotlib.pyplot as plt
from pylab import *


# 分别使用各个模型预测一些测试用例
def test_all_models_with_eg(model_path, review_paths, max_len, kw_max_len, dim):
    data_dict = preprocess_amazon_mobile_data(model_path, review_paths, max_len, kw_max_len, dim, lda=False, aug=False,
                                              LM=False)

    labels_file = open('trained_model/amazon_mobile/model_labels', 'w', encoding='utf-8')

    # 从字典中提取词语数量和词嵌入矩阵
    # 测试特定用例
    test_x = data_dict['test_x']
    test_y = data_dict['test_y']

    vocab_size = data_dict['vocab_size']
    embedding_matrix = data_dict['embed']

    # 加载所有模型
    # 多头注意力
    '''
    att = att300(max_len, vocab_size, dim, embedding_matrix)
    att_y = att.predict(test_x)
    att_labels = get_predicted_labels(att_y)
    labels_file.write('Att:')
    for label in att_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att.')
    '''
    # LSTM
    lstm_model = lstm(max_len, vocab_size, dim, embedding_matrix)
    lstm_y = lstm_model.predict(test_x)
    print(lstm_model.evaluate(test_x, test_y, verbose=2))
    lstm_labels = get_predicted_labels(lstm_y)
    labels_file.write('LSTM:')
    for label in lstm_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading LSTM.')

    # CNN
    cnn_model = orgcnn(max_len, vocab_size, dim, embedding_matrix)
    cnn_y = cnn_model.predict(test_x)
    print(cnn_model.evaluate(test_x, test_y, verbose=2))
    cnn_labels = get_predicted_labels(cnn_y)
    labels_file.write('CNN:')
    for label in cnn_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading CNN.')

    # CNN multiply LSTM
    att_model = cnned_lstm_multiply_lstm(max_len, vocab_size, dim, embedding_matrix)
    att_y = att_model.predict(test_x)
    print(att_model.evaluate(test_x, test_y, verbose=2))
    att_labels = get_predicted_labels(att_y)
    labels_file.write('Att:')
    for label in att_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    '''
    att_intern_out = temp_model.predict(test_x)
    for out in att_intern_out:
        for num in out:
            labels_file.write(str(num)+' ')
        labels_file.write('\n')
    '''
    print('Finish loading Att')
    '''
    # Att-LSTM
    att_lstm_m = att_lstm(max_len, vocab_size, dim, embedding_matrix)
    att_lstm_y = att_lstm_m.predict(test_x)
    att_lstm_labels = get_predicted_labels(att_lstm_y)
    labels_file.write('Att-LSTM:')
    for label in att_lstm_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att-LSTM.')

    # Att-LSTM-CNN
    att_lstm_cnn_m = att_lstm_cnn(max_len, vocab_size, dim, embedding_matrix)
    att_lstm_cnn_y = att_lstm_cnn_m.predict(test_x)
    att_lstm_cnn_labels = get_predicted_labels(att_lstm_cnn_y)
    labels_file.write('Att-LSTM-CNN:')
    for label in att_lstm_cnn_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att-LSTM-CNN.')

    # Att-LSTM-CNN(LDA)
    att_lstm_ldacnn_m = att_lstm_ldacnn(max_len, vocab_size, dim, embedding_matrix, kw_max_len)
    att_lstm_ldacnn_y = att_lstm_ldacnn_m.predict([test_x, test_kw_x])
    att_lstm_ldacnn_labels = get_predicted_labels(att_lstm_ldacnn_y)
    labels_file.write('Att-LSTM-LDACNN:')
    for label in att_lstm_ldacnn_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att-LSTM-LDACNN.')

    # Att-LSTM-BiCNN
    att300_lstm_bicnn = att300_lstm_bicnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix)
    att300_lstm_bicnn_y = att300_lstm_bicnn.predict([test_x, test_kw_x])
    att300_lstm_bicnn_labels = get_predicted_labels(att300_lstm_bicnn_y)
    labels_file.write('Att-LSTM-BiCNN:')
    for label in att300_lstm_bicnn_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att-LSTM-BiCNN.')

    # Att-2LSTM-CNN
    att300_2lstm_cnn = att300_bilstm_cnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix)
    att300_2lstm_cnn_y = att300_2lstm_cnn.predict([test_x, test_kw_x])
    att300_2lstm_cnn_labels = get_predicted_labels(att300_2lstm_cnn_y)
    labels_file.write('Att-2LSTM-CNN:')
    for label in att300_2lstm_cnn_labels:
        labels_file.write(str(label) + ' ')
    labels_file.write('\n')
    print('Finish loading Att-2LSTM-CNN.')
    '''
    print()


# 输出网络结构的中间结果进行分析
def output_intern_res(model_path, review_paths, max_len, kw_max_len, dim):
    data_dict = preprocess_amazon_mobile_data(model_path, review_paths, max_len, dim)

    eg_x = data_dict['eg_x']
    eg_y = data_dict['eg_y']

    vocab_size = data_dict['vocab_size']
    embedding_matrix = data_dict['embed']

    f = open('debug_data/amazon_mobile_att_out', 'w', encoding='utf-8')

    att_model = get_att_model(max_len, vocab_size, dim, embedding_matrix)
    att_out = att_model.predict(eg_x)
    for i in range(0, len(eg_x)):
        out3 = att_out[i]
        out3_max = select_max(out3)
        write_vec_to_file(f, out3_max)
        f.write('\n')


def select_max(matrix):
    max_vector = []
    for vec in matrix:
        max = vec[0]
        for num in vec:
            if num > max:
                max = num
        max_vector.append(max)
    return max_vector


def write_vec_to_file(file, vec):
    for num in vec:
        file.write(str(num)+' ')
    file.write('\n')


##########################################################################

# 探究多头注意力中的头数对结果的影响
def test_head_effect(model_path, review_paths, max_len, kw_max_len, dim):
    '''
    data_dict = preprocess_amazon_mobile_data(model_path, review_paths, max_len, kw_max_len, dim, lda=False)

    # 从字典中提取词语数量和词嵌入矩阵
    # 测试特定用例
    test_x = data_dict['test_x']
    test_y = data_dict['test_y']
    #test_kw_x = data_dict['test_kw_x']

    vocab_size = data_dict['vocab_size']
    embedding_matrix = data_dict['embed']

    # 加载不同头数的多头注意力模型
    att = att300(max_len, vocab_size, dim, embedding_matrix)
    att300_y = att.predict(test_x)
    att300_labels = get_predicted_labels(att300_y)

    att = att600(max_len, vocab_size, dim, embedding_matrix)
    att600_y = att.predict(test_x)
    att600_labels = get_predicted_labels(att600_y)

    att = att1200(max_len, vocab_size, dim, embedding_matrix)
    att1200_y = att.predict(test_x)
    att1200_labels = get_predicted_labels(att1200_y)
    '''
    att300_labels = []
    att600_labels = []
    att1200_labels = []
    with open('debug_data/attention_labels', 'r', encoding='utf-8') as f:
        for line in f:
            if len(att300_labels) < 2500:
                att300_labels.append(line.replace('\n', '').strip())
            elif len(att600_labels) < 2500:
                att600_labels.append(line.replace('\n', '').strip())
            elif len(att1200_labels) < 2500:
                att1200_labels.append(line.replace('\n', '').strip())

    sentences, labels = read_test_data()
    wrong_sents = []
    wrong_labels_list = []
    print('Picking out wrong samples...')
    for i in range(0, len(labels)):
        if str(att300_labels[i]) == str(labels[i]) and str(att600_labels[i]) == str(labels[i]) and \
                str(att1200_labels[i]) == str(labels[i]):
            continue
        else:
            wrong_sents.append(sentences[i])
            wrong_labels = []
            wrong_labels.append(str(labels[i]))
            wrong_labels.append(str(att300_labels[i]))
            wrong_labels.append(str(att600_labels[i]))
            wrong_labels.append(str(att1200_labels[i]))
            wrong_labels_list.append(wrong_labels)


    print('Writing to xlsx...')
    wb = xl.Workbook()
    ws = wb.create_sheet('Attention wrong samples')
    for i in range(0, len(wrong_sents)):
        sent = wrong_sents[i]
        wrong_labels = wrong_labels_list[i]
        ws['A%d' % (i + 1)] = sent
        ws['B%d' % (i + 1)] = wrong_labels[0]
        if wrong_labels[0] != wrong_labels[1]:
            ws['C%d' % (i + 1)] = wrong_labels[1]+'***'
        else:
            ws['C%d' % (i + 1)] = wrong_labels[1]
        if wrong_labels[0] != wrong_labels[2]:
            ws['D%d' % (i + 1)] = wrong_labels[2]+'***'
        else:
            ws['D%d' % (i + 1)] = wrong_labels[2]
        if wrong_labels[0] != wrong_labels[3]:
            ws['E%d' % (i + 1)] = wrong_labels[3]+'***'
        else:
            ws['E%d' % (i + 1)] = wrong_labels[3]
    wb.save('debug_data/attention_wrong_sent.xlsx')


def att300(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    dense = Dense(300, activation='relu')(att_pool_dp)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Loading multi-head attention...")
    model.load_weights('trained_model/amazon_mobile/att300_w.h5')
    return model


def att600(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(20, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    dense = Dense(500, activation='relu')(att_pool_dp)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/att600_w.h5')
    return model


def att1200(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(40, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    dense = Dense(800, activation='relu')(att_pool_dp)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/att1200_w.h5')
    return model


def lstm(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    lstm = Bidirectional(LSTM(max_len, dropout=0.3))(embed_dp)

    dense = Dense(2 * max_len, activation='relu')(lstm)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/lstm_w.h5')
    return model


def cnned_lstm(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=300, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp)
    conv_lstm = Bidirectional(LSTM(max_len, dropout=0.3, ))(conv)

    dense = Dense(2 * max_len, activation='relu')(conv_lstm)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_food/cnned_lstm_w.h5')
    return model


def cnned_att_and_cnned_lstm(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=300, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_lstm = Bidirectional(LSTM(max_len,
                                   dropout=0.3,
                                   return_sequences=False,
                                   return_state=False))(conv)

    att = Attention(20, 30)([conv, conv, conv])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    merged_vector = concatenate([conv_lstm, att_pool_dp], axis=-1)

    dense = Dense(600, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print(model.summary())
    print("Start training...")
    model.save_weights('trained_model/amazon_mobile/cnned_lstm_and_cnned_att_w.h5')
    return model


def cnn_multiply_lstm(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=max_len * 2, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp)
    lstm = Bidirectional(LSTM(max_len, dropout=0.3, return_sequences=True, return_state=False))(embed_dp)
    att = Multiply()([lstm, conv])
    pool = GlobalMaxPool1D()(att)

    dense = Dense(2 * max_len, activation='relu')(pool)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.load_weights('trained_model/amazon_mobile/cnn_multiply_lstm_w.h5')


def ldacnn():
    print("Loading CNN based on LDA...")
    model = load_model('trained_model/amazon_mobile/ldacnn_ap.h5')
    return model


def orgcnn(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_pool = GlobalMaxPool1D()(conv)
    conv_pool_dp = Dropout(0.3)(conv_pool)

    dense = Dense(150, activation='relu')(conv_pool_dp)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print('Loading CNN...')
    model.load_weights('trained_model/amazon_mobile/cnn_w.h5')
    return model


def lstm_orgcnn():
    print("Loading LSTM-CNN...")
    model = load_model('trained_model/amazon_mobile/lstm_orgcnn.h5')
    return model


def lstm_ldacnn():
    print("Loading LSTM-LDACNN...")
    model = load_model('trained_model/amazon_mobile/lstm_ldacnn.h5')
    return model


def lstm_bicnn():
    print("Loading LSTM-BiCNN...")
    model = load_model('trained_model/amazon_mobile/lstm_orgcnn_ldacnn.h5')
    return model


def att_lstm(max_len, vocab_size, dim, embedding_matrix, kw_max_len=None):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    merged_vector = concatenate([att_pool_dp, lstm_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print('Loading att-lstm...')
    model.load_weights('trained_model/amazon_mobile/att_lstm_w.h5')
    return model


def att600_lstm(max_len, vocab_size, dim, embedding_matrix, kw_max_len=None):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(20, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    merged_vector = concatenate([att_pool_dp, lstm_dp], axis=-1)

    dense = Dense(600, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/att600_lstm_w.h5')
    return model


def att_lstm_cnn(max_len, vocab_size, dim, embedding_matrix, kw_max_len=None):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn], axis=-1)

    dense = Dense(450, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print('Loading att-lstm-cnn model...')
    model.load_weights('trained_model/amazon_mobile/att_lstm_cnn_w.h5')
    return model


def att600_lstm_cnn(max_len, vocab_size, dim, embedding_matrix, kw_max_len=None):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att = Attention(20, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn], axis=-1)

    dense = Dense(700, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/att600_lstm_cnn_w.h5')


def cnned_lstm_att600(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=300, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp)
    conv_lstm = Bidirectional(LSTM(max_len,
                                   dropout=0.3,
                                   return_sequences=True,
                                   return_state=False))(conv)

    att = Attention(20, 30)([conv_lstm, conv_lstm, conv_lstm])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    dense = Dense(600, activation='relu')(att_pool_dp)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/cnned_lstm_multihead_w.h5')
    return model


def att_lstm_ldacnn(max_len, vocab_size, dim, embedding_matrix, kw_max_len=None):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([att_pool_dp, lstm_dp, kw_pool_dp], axis=-1)

    dense = Dense(450, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Loading Att-LSTM-LDACNN...")
    model.load_weights('trained_model/amazon_mobile/att300_lstm_ldacnn_w.h5')
    return model


def att_lstm_bicnn(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(3, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_pool = GlobalMaxPool1D()(conv)
    conv_pool_dp = Dropout(0.3)(conv_pool)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([lstm_dp, Add()([conv_pool_dp, kw_pool_dp]), att_pool_dp], axis=-1)

    dence = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.3)(dence)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Load att-LSTM-BiCNN...")
    model.load_weights('trained_model/amazon_mobile/all_with_neg_words_w.h5')
    return model


def att_lstm_bicnn_full_lda_add(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(3, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_pool = GlobalMaxPool1D()(conv)
    conv_pool_dp = Dropout(0.3)(conv_pool)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([lstm_dp, Add()([conv_pool_dp, kw_pool_dp]), att_pool_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/all_full_lda_w.h5')
    return model


def att_lstm_bicnn_full_lda_concat(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(3, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_pool = GlobalMaxPool1D()(conv)
    conv_pool_dp = Dropout(0.3)(conv_pool)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([lstm_dp, conv_pool_dp, kw_pool_dp, att_pool_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    model.load_weights('trained_model/amazon_mobile/all_full_lda_concat_w.h5')
    return model


# 经过两次过滤的LDA进行CNN处理
def att_lstm_bicnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(3, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([lstm_dp, conv_bn, kw_pool_dp, att_pool_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Load all with 2 filter of LDA...")
    model.load_weights('trained_model/amazon_mobile/all_2filter_lda_w.h5')
    return model


def att_bilstm_cnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(3, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    kw_lstm = Bidirectional(LSTM(kw_max_len))(kw_embed_dp)
    kw_lstm_dp = Dropout(0.3)(kw_lstm)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn, kw_lstm_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print(model.summary())
    print("Start training...")
    model.load_weights('trained_model/amazon_mobile/att_2bilstm_cnn_w.h5')
    return model


def att500_bilstm_cnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(10, 50)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    kw_lstm = Bidirectional(LSTM(kw_max_len))(kw_embed_dp)
    kw_lstm_dp = Dropout(0.3)(kw_lstm)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn, kw_lstm_dp], axis=-1)

    dense = Dense(300, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Loading att500 bilstm cnn...")
    model.load_weights('trained_model/amazon_mobile/att500_2bilstm_cnn_w.h5')
    return model


def att300_lstm_bicnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    kw_conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(kw_embed_dp)
    kw_pool = GlobalAveragePooling1D()(kw_conv)
    kw_pool_dp = Dropout(0.3)(kw_pool)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn, kw_pool_dp], axis=-1)

    dense = Dense(600, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Loading att300-lstm-bicnn...")
    model.load_weights('trained_model/amazon_mobile/att300_lstm_bicnn_w.h5')
    return model


def att300_bilstm_cnn_with_filter_2times_lda(max_len, kw_max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    kw_input = Input(shape=(kw_max_len,), dtype='int32')  # LDA文本
    kw_embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=kw_max_len, trainable=False)(
        kw_input)
    kw_embed_dp = Dropout(0.3)(kw_embed)

    att = Attention(10, 30)([embed_dp, embed_dp, embed_dp])
    att_pool = GlobalAveragePooling1D()(att)
    att_pool_dp = Dropout(0.3)(att_pool)

    lstm = Bidirectional(LSTM(max_len))(embed_dp)
    lstm_dp = Dropout(0.3)(lstm)

    conv = Conv1D(filters=150, kernel_size=3, padding='valid', activation='relu', strides=1)(embed_dp)
    conv_dp = Dropout(0.3)(conv)
    conv_pool_bn = BatchNormalization()(conv_dp)
    conv2 = Conv1D(filters=50, kernel_size=3, padding='valid', activation='relu', strides=1)(conv_pool_bn)
    conv2_dp = Dropout(0.3)(conv2)
    conv2_pool_bn = BatchNormalization()(conv2_dp)
    flatten = Flatten()(conv2_pool_bn)
    conv_dense = Dense(100, activation='relu')(flatten)
    conv_dense_dp = Dropout(0.3)(conv_dense)
    conv_bn = BatchNormalization()(conv_dense_dp)

    kw_lstm = Bidirectional(LSTM(kw_max_len))(kw_embed_dp)
    kw_lstm_dp = Dropout(0.3)(kw_lstm)

    merged_vector = concatenate([att_pool_dp, lstm_dp, conv_bn, kw_lstm_dp], axis=-1)

    dense = Dense(550, activation='relu')(merged_vector)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=[input, kw_input], outputs=predictions)
    model.compile(loss='sparse_categorical_crossentropy',
                  optimizer='adam',
                  metrics=['accuracy'])
    print("Loading att300-2lstm-cnn...")
    model.load_weights('trained_model/amazon_mobile/att300_2lstm_cnn_w.h5')
    return model


def get_predicted_labels(predict_y):
    labels = []
    for y in predict_y:
        max_value = max(y)
        if max_value == y[0]:
            labels.append(1)
        elif max_value == y[1]:
            labels.append(2)
        elif max_value == y[2]:
            labels.append(3)
        elif max_value == y[3]:
            labels.append(4)
        elif max_value == y[4]:
            labels.append(5)
    return labels


##################################################
# 输出模型中间向量
def get_lstm_model(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    # LSTM 输出
    lstm = Bidirectional(LSTM(max_len, dropout=0.3,
                              return_sequences=True))(embed_dp)

    lstm_model = Model(inputs=input, outputs=lstm)
    lstm_model.load_weights('trained_model/amazon_mobile/lstm_w.h5', by_name=True)

    return lstm_model


def get_cnn_lstm_model(max_len, vocab_size, dim, embedding_matrix):
    # CNN-LSTM输出
    input2 = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed2 = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input2)
    embed_dp2 = Dropout(0.3)(embed2)

    conv = Conv1D(filters=300, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp2)
    conv_lstm = Bidirectional(LSTM(max_len, dropout=0.3, return_sequences=True))(conv)

    cnn_lstm_model = Model(inputs=input2, outputs=conv_lstm)
    cnn_lstm_model.load_weights('trained_model/amazon_mobile/cnn_lstm_w.h5', by_name=True)

    return cnn_lstm_model


def get_att_model(max_len, vocab_size, dim, embedding_matrix):
    # CNN-Multiply-LSTM
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=300, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp)
    conv_lstm = Bidirectional(LSTM(max_len,
                                   dropout=0.3,
                                   return_sequences=True,
                                   return_state=False))(conv)

    lstm = Bidirectional(LSTM(max_len,
                              dropout=0.3,
                              return_sequences=True,
                              return_state=False))(embed_dp)

    att = Multiply()([lstm, conv_lstm])
    pool = GlobalMaxPool1D()(att)

    dense = Dense(2 * max_len, activation='relu')(pool)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    att_model = Model(inputs=input, outputs=att)
    att_model.load_weights('trained_model/amazon_mobile/cnned_lstm_multiply_lstm_w.h5', by_name=True)
    return att_model


# 绘制网络结构中间输出的折线图
def draw_output():
    lstm_outputs = []
    cnn_lstm_outputs = []
    att_outputs = []
    with open('debug_data/amazon_mobile_lstm_out', 'r', encoding='utf-8') as f:
        for line in f:
            out = []
            array = line.replace('\n', '').split()
            if len(array) == 0:
                continue
            for num in array:
                out.append(float(num))
            lstm_outputs.append(out)
    with open('debug_data/amazon_mobile_cnn_lstm_out', 'r', encoding='utf-8') as f:
        for line in f:
            out = []
            array = line.replace('\n', '').split()
            if len(array) == 0:
                continue
            for num in array:
                out.append(float(num))
            cnn_lstm_outputs.append(out)
    with open('debug_data/amazon_mobile_att_out', 'r', encoding='utf-8') as f:
        for line in f:
            out = []
            array = line.replace('\n', '').split()
            if len(array) == 0:
                continue
            for num in array:
                out.append(float(num))
            att_outputs.append(out)

    x = list(range(80))
    y1 = lstm_outputs[10]
    y2 = cnn_lstm_outputs[10]
    y3 = att_outputs[8]

    #plt.plot(x, y1, label='lstm')
    #plt.plot(x, y2, label='cnn-lstm')
    plt.plot(x, y3, label='att')
    plt.legend()
    plt.xlabel('dim')
    plt.ylabel('value')
    plt.show()


########################################################################
# 几篇对比论文的结构
# CNN-RNN based attention model
def CRAN(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv1 = Conv1D(filters=max_len * 2, kernel_size=3, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    conv2 = Conv1D(filters=max_len * 2, kernel_size=4, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    conv3 = Conv1D(filters=max_len * 2, kernel_size=5, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    pool1 = GlobalMaxPool1D()(conv1)
    pool2 = GlobalMaxPool1D()(conv2)
    pool3 = GlobalMaxPool1D()(conv3)
    pool_cat = concatenate([pool1, pool2, pool3], axis=-1)
    pool_dense = Dense(max_len * 2, activation='relu')(pool_cat)

    lstm = Bidirectional(LSTM(max_len, dropout=0.3))(embed_dp)
    lstm_dense = Dense(max_len * 2, activation='relu')(lstm)

    att = Multiply()([pool_dense, lstm_dense])
    predictions = Dense(5, activation='softmax')(att)

    model = Model(inputs=input, outputs=predictions)
    model.load_weights('trained_model/amazon_mobile/CRAN_w.h5')
    return model


# Convolutional attention model
def CAM(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv1 = Conv1D(filters=max_len * 2, kernel_size=2, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    conv2 = Conv1D(filters=max_len * 2, kernel_size=3, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    conv3 = Conv1D(filters=max_len * 2, kernel_size=4, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)
    conv4 = Conv1D(filters=max_len * 2, kernel_size=5, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)

    avg = Average()([conv1, conv2, conv3, conv4])

    lstm = Bidirectional(LSTM(max_len, dropout=0.3, return_sequences=True, return_state=False))(embed_dp)

    att = Dot(axes=-1)([avg, lstm])
    att_avg = GlobalAveragePooling1D()(att)
    predictions = Dense(5, activation='softmax')(att_avg)

    model = Model(inputs=input, outputs=predictions)
    model.save_weights('trained_model/amazon_food/CAM_w.h5')
    return model


# double attention
def DA(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    att_lstm = model_lstm_with_self_att(embed_dp, max_len)

    conv = Conv1D(filters=max_len * 2, kernel_size=3, padding='same', activation='relu', strides=1, use_bias=True)(
        embed_dp)

    cossim = Dot(axes=-1, normalize=True)([att_lstm, conv])
    final_rep = Dot(axes=1)([cossim, conv])
    predictions = Dense(5, activation='softmax')(final_rep)

    model = Model(inputs=input, outputs=predictions)
    model.save_weights('trained_model/amazon_mobile/DA_w.h5')


# 返回LSTM与自注意力结合的结果
def model_lstm_with_self_att(embed_dp, max_len):
    hidden_states = embed_dp
    hidden_states = Bidirectional(LSTM(max_len,
                                           dropout=0.3,
                                           return_sequences=True,
                                           return_state=False))(hidden_states)

    # Attention mechanism
    attention = Conv1D(filters=max_len, kernel_size=1, activation='tanh', padding='same', use_bias=True,
                       kernel_initializer='glorot_uniform', bias_initializer='zeros', name="attention_layer1")(hidden_states)
    attention = Conv1D(filters=max_len, kernel_size=1, activation='linear', padding='same',use_bias=True,
                       kernel_initializer='glorot_uniform', bias_initializer='zeros',
                       name="attention_layer2")(attention)
    attention = Lambda(lambda x: softmax(x, axis=1), name="attention_vector")(attention)

    # Apply attention weights
    weighted_sequence_embedding = Dot(axes=[1, 1], normalize=False, name="weighted_sequence_embedding")(
        [attention, hidden_states])

    # Add and normalize to obtain final sequence embedding
    sequence_embedding = Lambda(lambda x: K.l2_normalize(K.sum(x, axis=1)))(weighted_sequence_embedding)

    return sequence_embedding


# Combination of CNN and GRU
def CCG(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    gru = Bidirectional(GRU(max_len, dropout=0.3, return_sequences=True, return_state=False))(embed_dp)
    conv = Conv1D(filters=max_len * 2, kernel_size=3, strides=1, padding='same', use_bias=True)(embed_dp)

    s = Multiply()([gru, conv])
    g = Multiply()([gru, s])
    pool = GlobalMaxPool1D()(g)

    predictions = Dense(5, activation='softmax')(pool)

    model = Model(inputs=input, outputs=predictions)
    model.load_weights('trained_model/amazon_mobile/CCG_w.h5')


# ours
def cnned_lstm_multiply_lstm(max_len, vocab_size, dim, embedding_matrix):
    input = Input(shape=(max_len,), dtype='int32')  # 原始文本
    embed = Embedding(vocab_size, dim, weights=[embedding_matrix], input_length=max_len, trainable=False)(input)
    embed_dp = Dropout(0.3)(embed)

    conv = Conv1D(filters=300, kernel_size=3, padding='same', activation='relu', strides=1)(embed_dp)
    conv_lstm = Bidirectional(LSTM(max_len,
                                   dropout=0.3,
                                   return_sequences=True,
                                   return_state=False))(conv)

    lstm = Bidirectional(LSTM(max_len,
                              dropout=0.3,
                              return_sequences=True,
                              return_state=False))(embed_dp)

    att = Multiply()([lstm, conv_lstm])
    pool = GlobalMaxPool1D()(att)

    dense = Dense(2*max_len, activation='relu')(pool)
    dp = Dropout(0.32)(dense)
    predictions = Dense(5, activation='softmax')(dp)

    model = Model(inputs=input, outputs=predictions)
    print('Loading Att...')
    model.load_weights('trained_model/amazon_mobile/cnned_lstm_multiply_lstm_w.h5')
    # 加载部分模型查看中间的输出向量
    return model #temp_model


# 预训练好的词向量模型
glove_300d_path = "yelp_glove/glove.840B.300d.txt"
glove_100d_path = "yelp_glove/glove.6B.100d.txt"
ic_self_trained_model_path = "output_result/up_down_stream.vector"
fast_text_wiki_path = "fast_text_vec/wiki-news-300d-1M.vec"
fast_text_crawl_path = "fast_text_vec/crawl-300d-2M.vec"


# Yelp训练和测试样本，十个样本文件
rv_path1 = "yelp_review_1/yelp_review1"
rv_path2 = "yelp_review_1/yelp_review2"
rv_path3 = "yelp_review_1/yelp_review3"
rv_path4 = "yelp_review_1/yelp_review4"
rv_path5 = "yelp_review_1/yelp_review5"
rv_path6 = "yelp_review_1/yelp_review6"
rv_path7 = "yelp_review_1/yelp_review7"
rv_path8 = "yelp_review_1/yelp_review8"
rv_path9 = "yelp_review_1/yelp_review9"
rv_path10 = "yelp_review_1/yelp_review10"

# 产业链数据
ic_path = "training_data/seg_sent_with_label2"
ic_paths = []
ic_paths.append(ic_path)

# IMDB影评训练、测试数据
imdb_train_path = 'Imdb_reviews/train_data'
imdb_test_path = 'Imdb_reviews/test_data'

# Amazon食品评论数据
amazon_reviews_path = 'Amazon_reviews/food_reviews.csv'
food_reviews = []
food_reviews.append('Amazon_reviews/train_double')
food_reviews.append('Amazon_reviews/test_double')

# Amazon电子设备评论数据
mobile_reviews = []
mobile_reviews.append('Amazon_mobile_reviews/train')
mobile_reviews.append('Amazon_mobile_reviews/test')
mobile_reviews.append('Amazon_mobile_reviews/dev')

# 文本蕴含数据
etm_path = "entailment_data/gov_data.txt"

# Yelp评论数据
yelp_paths = []
yelp_paths.append('yelp_new_reviews/train_double')
yelp_paths.append('yelp_new_reviews/test_double')

draw_output()
#output_intern_res(glove_300d_path, mobile_reviews, max_len=80, kw_max_len=25, dim=300)
#test_all_models_with_eg(glove_300d_path, mobile_reviews, max_len=80, kw_max_len=25, dim=300)
#test_head_effect(glove_300d_path, mobile_reviews, max_len=80, kw_max_len=25, dim=300)
